# reporting/excel_export.py
"""Generate analytical Excel export from the PianoTempi history tables."""
import logging
from datetime import date
from io import BytesIO
from typing import List, Dict

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from data_sources.db_history import (
    get_plan_history_for_date, get_prod_history_for_date, get_cycles_master,
)

logger = logging.getLogger("PianoTempi")

HEADER_FILL = PatternFill(start_color="3880FF", end_color="3880FF", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
NUM_COL_FILL = PatternFill(start_color="F2F8FF", end_color="F2F8FF", fill_type="solid")


def _autosize(ws):
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        for cell in ws[col_letter]:
            v = "" if cell.value is None else str(cell.value)
            if len(v) > max_len:
                max_len = len(v)
        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)


def _write_header(ws, headers: List[str]) -> None:
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _build_summary_sheet(ws, plan_rows: List[Dict], prod_rows: List[Dict]) -> None:
    """Aggregate plan + prod by phase. One row per phase."""
    ws.title = "Summary"
    headers = ["Phase", "Planned hours", "Produced hours", "Coverage %", "Delta hours"]
    _write_header(ws, headers)

    by_phase: Dict[str, Dict[str, float]] = {}
    for r in plan_rows:
        p = r["phase_name"]
        d = by_phase.setdefault(p, {"plan": 0.0, "prod": 0.0})
        d["plan"] += float(r["planned_hours"] or 0.0)
    for r in prod_rows:
        p = r["phase_name"]
        d = by_phase.setdefault(p, {"plan": 0.0, "prod": 0.0})
        d["prod"] += float(r["produced_hours"] or 0.0)

    row_idx = 2
    tot_plan = tot_prod = 0.0
    for phase, agg in sorted(by_phase.items()):
        cov = (agg["prod"] / agg["plan"] * 100.0) if agg["plan"] > 0 else 0.0
        delta = agg["prod"] - agg["plan"]
        ws.cell(row=row_idx, column=1, value=phase)
        ws.cell(row=row_idx, column=2, value=round(agg["plan"], 2))
        ws.cell(row=row_idx, column=3, value=round(agg["prod"], 2))
        ws.cell(row=row_idx, column=4, value=round(cov, 2))
        ws.cell(row=row_idx, column=5, value=round(delta, 2))
        tot_plan += agg["plan"]; tot_prod += agg["prod"]
        row_idx += 1

    # Total row
    ws.cell(row=row_idx, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=row_idx, column=2, value=round(tot_plan, 2)).font = Font(bold=True)
    ws.cell(row=row_idx, column=3, value=round(tot_prod, 2)).font = Font(bold=True)
    cov = (tot_prod / tot_plan * 100.0) if tot_plan > 0 else 0.0
    ws.cell(row=row_idx, column=4, value=round(cov, 2)).font = Font(bold=True)
    ws.cell(row=row_idx, column=5, value=round(tot_prod - tot_plan, 2)).font = Font(bold=True)

    _autosize(ws)


def _build_plan_detail_sheet(ws, plan_rows: List[Dict]) -> None:
    ws.title = "Plan detail"
    headers = ["Plan date", "Phase", "Order number", "Product code",
               "Planned qty", "Cycle min", "Planned minutes", "Planned hours"]
    _write_header(ws, headers)
    row_idx = 2
    for r in plan_rows:
        ws.cell(row=row_idx, column=1, value=r["plan_date"].isoformat() if r.get("plan_date") else "")
        ws.cell(row=row_idx, column=2, value=r.get("phase_name", ""))
        ws.cell(row=row_idx, column=3, value=r.get("order_number", ""))
        ws.cell(row=row_idx, column=4, value=r.get("product_code", ""))
        ws.cell(row=row_idx, column=5, value=int(r.get("planned_qty") or 0))
        ws.cell(row=row_idx, column=6, value=float(r.get("cycle_time_minutes") or 0))
        ws.cell(row=row_idx, column=7, value=float(r.get("planned_minutes") or 0))
        ws.cell(row=row_idx, column=8, value=float(r.get("planned_hours") or 0))
        row_idx += 1
    _autosize(ws)


def _build_prod_detail_sheet(ws, prod_rows: List[Dict]) -> None:
    ws.title = "Production detail"
    headers = ["Prod date", "Phase", "Shift", "Order number", "Product code",
               "Trace ID", "Produced qty", "Cycle min", "Produced minutes", "Produced hours"]
    _write_header(ws, headers)
    row_idx = 2
    for r in prod_rows:
        ws.cell(row=row_idx, column=1, value=r["prod_date"].isoformat() if r.get("prod_date") else "")
        ws.cell(row=row_idx, column=2, value=r.get("phase_name", ""))
        ws.cell(row=row_idx, column=3, value=r.get("shift_code", ""))
        ws.cell(row=row_idx, column=4, value=r.get("order_number", ""))
        ws.cell(row=row_idx, column=5, value=r.get("product_code", ""))
        ws.cell(row=row_idx, column=6, value=r.get("traceability_phase_id"))
        ws.cell(row=row_idx, column=7, value=int(r.get("produced_qty") or 0))
        ws.cell(row=row_idx, column=8, value=float(r.get("cycle_time_minutes") or 0))
        ws.cell(row=row_idx, column=9, value=float(r.get("produced_minutes") or 0))
        ws.cell(row=row_idx, column=10, value=float(r.get("produced_hours") or 0))
        row_idx += 1
    _autosize(ws)


def _build_comparison_sheet(ws, plan_rows: List[Dict], prod_rows: List[Dict]) -> None:
    """Joined view: per (order, phase) -> planned + produced (sum across shifts) + delta."""
    ws.title = "Comparison"
    headers = ["Phase", "Order number", "Product code",
               "Planned qty", "Planned hours",
               "Produced qty", "Produced hours",
               "Coverage %", "Delta hours"]
    _write_header(ws, headers)

    # Aggregate plan by (order, phase)
    plan_agg: Dict[tuple, Dict] = {}
    for r in plan_rows:
        key = (r["order_number"], r["phase_name"])
        d = plan_agg.setdefault(key, {"product": r.get("product_code", ""), "qty": 0, "hours": 0.0})
        d["qty"] += int(r.get("planned_qty") or 0)
        d["hours"] += float(r.get("planned_hours") or 0)

    # Aggregate prod by (order, phase) summed across shifts
    prod_agg: Dict[tuple, Dict] = {}
    for r in prod_rows:
        key = (r["order_number"], r["phase_name"])
        d = prod_agg.setdefault(key, {"product": r.get("product_code", ""), "qty": 0, "hours": 0.0})
        d["qty"] += int(r.get("produced_qty") or 0)
        d["hours"] += float(r.get("produced_hours") or 0)

    all_keys = sorted(set(plan_agg.keys()) | set(prod_agg.keys()),
                      key=lambda k: (k[1], k[0]))  # by phase then order
    row_idx = 2
    for (order_num, phase) in all_keys:
        p = plan_agg.get((order_num, phase), {"product": "", "qty": 0, "hours": 0.0})
        d = prod_agg.get((order_num, phase), {"product": "", "qty": 0, "hours": 0.0})
        product = p["product"] or d["product"]
        cov = (d["hours"] / p["hours"] * 100.0) if p["hours"] > 0 else 0.0
        delta = d["hours"] - p["hours"]
        ws.cell(row=row_idx, column=1, value=phase)
        ws.cell(row=row_idx, column=2, value=order_num)
        ws.cell(row=row_idx, column=3, value=product)
        ws.cell(row=row_idx, column=4, value=p["qty"])
        ws.cell(row=row_idx, column=5, value=round(p["hours"], 2))
        ws.cell(row=row_idx, column=6, value=d["qty"])
        ws.cell(row=row_idx, column=7, value=round(d["hours"], 2))
        ws.cell(row=row_idx, column=8, value=round(cov, 2))
        ws.cell(row=row_idx, column=9, value=round(delta, 2))
        row_idx += 1
    _autosize(ws)


def build_excel_for_date(conn, target_date: date) -> bytes:
    """Build a 4-sheet Excel report for the given date and return the bytes."""
    plan_rows = get_plan_history_for_date(conn, target_date)
    prod_rows = get_prod_history_for_date(conn, target_date)

    wb = openpyxl.Workbook()

    # First sheet (default): Summary
    ws_summary = wb.active
    _build_summary_sheet(ws_summary, plan_rows, prod_rows)

    ws_plan = wb.create_sheet("Plan detail")
    _build_plan_detail_sheet(ws_plan, plan_rows)

    ws_prod = wb.create_sheet("Production detail")
    _build_prod_detail_sheet(ws_prod, prod_rows)

    ws_cmp = wb.create_sheet("Comparison")
    _build_comparison_sheet(ws_cmp, plan_rows, prod_rows)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    logger.info(
        "Excel export built for %s: %d plan rows + %d prod rows",
        target_date, len(plan_rows), len(prod_rows),
    )
    return buf.read()
