# tests/test_excel_export.py
from datetime import date
from io import BytesIO
from unittest.mock import patch
import openpyxl
from reporting.excel_export import build_excel_for_date


@patch("reporting.excel_export.get_plan_history_for_date")
@patch("reporting.excel_export.get_prod_history_for_date")
def test_build_excel_has_4_sheets(mock_prod, mock_plan):
    mock_plan.return_value = [
        {"plan_date": date(2026, 5, 7), "order_number": "ORD1",
         "phase_name": "ASSEMBLY", "product_code": "P1",
         "planned_qty": 100, "cycle_time_minutes": 2.5,
         "planned_minutes": 250.0, "planned_hours": 4.17},
    ]
    mock_prod.return_value = [
        {"prod_date": date(2026, 5, 7), "order_number": "ORD1",
         "phase_name": "ASSEMBLY", "shift_code": "T1", "product_code": "P1",
         "traceability_phase_id": 110, "produced_qty": 30,
         "cycle_time_minutes": 2.5, "produced_minutes": 75.0, "produced_hours": 1.25},
    ]
    xlsx_bytes = build_excel_for_date(conn=None, target_date=date(2026, 5, 7))
    wb = openpyxl.load_workbook(BytesIO(xlsx_bytes))
    assert wb.sheetnames == ["Summary", "Plan detail", "Production detail", "Comparison"]
    # Summary should have ASSEMBLY row + TOTAL row
    ws = wb["Summary"]
    cell_a2 = ws["A2"].value
    assert cell_a2 == "ASSEMBLY"


@patch("reporting.excel_export.get_plan_history_for_date")
@patch("reporting.excel_export.get_prod_history_for_date")
def test_build_excel_empty_data(mock_prod, mock_plan):
    """An empty day still produces a valid xlsx."""
    mock_plan.return_value = []
    mock_prod.return_value = []
    xlsx_bytes = build_excel_for_date(conn=None, target_date=date(2026, 5, 7))
    wb = openpyxl.load_workbook(BytesIO(xlsx_bytes))
    assert wb.sheetnames == ["Summary", "Plan detail", "Production detail", "Comparison"]
