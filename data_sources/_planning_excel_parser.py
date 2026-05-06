# data_sources/_planning_excel_parser.py
"""Internal: copy of PlanRespect's planning Excel parser, adapted for PianoTempi."""
import logging
from dataclasses import dataclass
from datetime import date, datetime
from typing import List, Optional

import openpyxl

logger = logging.getLogger("PianoTempi")


@dataclass(frozen=True)
class _PlanRowRaw:
    order_number: str
    phase_name: str
    production_date: date
    planned_qty: int


def _parse_date_header(value) -> Optional[date]:
    """Convert a column header value to a date."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = str(value).strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y", "%d.%m.%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    logger.warning("date header not parsable: '%s'", s)
    return None


def _parse_qty(value) -> int:
    if value is None:
        return 0
    if isinstance(value, (int, float)):
        return int(value) if value == value else 0
    s = str(value).strip()
    if not s:
        return 0
    try:
        return int(float(s))
    except (ValueError, TypeError):
        return 0


def parse_last_phase(file_path: str, sheet_name: str = "PlanningMachine") -> List[_PlanRowRaw]:
    """Parse the planning sheet and return raw rows.

    Layout from PlanRespect:
      - column E (idx 4) = machine/phase name
      - column K (idx 10) = order number (with optional bullet prefix)
      - columns U+ (idx 20+) = date headers, cells are quantities
    """
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    except Exception as e:
        logger.error("cannot open planning xlsx %s: %s", file_path, e)
        return []

    if sheet_name not in wb.sheetnames:
        logger.error("sheet '%s' not in %s -- sheets: %s",
                     sheet_name, file_path, wb.sheetnames)
        wb.close()
        return []

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    wb.close()
    if not rows:
        return []

    header_row = rows[0]
    date_columns = {}
    for col_idx in range(20, len(header_row)):
        parsed = _parse_date_header(header_row[col_idx])
        if parsed:
            date_columns[col_idx] = parsed
    if not date_columns:
        logger.error("no valid date columns found from index 20 onwards")
        return []

    plan_rows: List[_PlanRowRaw] = []
    for row in rows[1:]:
        if len(row) < 21:
            continue
        raw_order = row[10]
        if raw_order is None:
            continue
        order_number = str(raw_order).lstrip("•").strip()
        if not order_number:
            continue
        raw_phase = row[4]
        if raw_phase is None:
            continue
        phase_name = str(raw_phase).strip()
        if not phase_name:
            continue
        for col_idx, prod_date in date_columns.items():
            if col_idx < len(row):
                qty = _parse_qty(row[col_idx])
                if qty > 0:
                    plan_rows.append(_PlanRowRaw(
                        order_number=order_number,
                        phase_name=phase_name,
                        production_date=prod_date,
                        planned_qty=qty,
                    ))
    logger.info("planning parsed: %d rows from %s", len(plan_rows), file_path)
    return plan_rows
