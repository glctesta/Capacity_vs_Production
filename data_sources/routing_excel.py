"""Read the routing Excel from T:\\D365 routing data.
Cells with 'x' (any case) are skipped. Non-numeric cells logged + skipped.
"""
import logging
import os
from datetime import datetime
from typing import Dict, Optional, Tuple

import openpyxl

logger = logging.getLogger("PianoTempi")


def find_latest_routing_file(folder: str) -> Optional[str]:
    """Return absolute path of most recent .xlsx in folder, ignoring temp files."""
    if not os.path.isdir(folder):
        logger.error("routing folder not accessible: %s", folder)
        return None
    candidates = []
    for name in os.listdir(folder):
        if name.startswith("~$"):
            continue
        if not name.lower().endswith(".xlsx"):
            continue
        full = os.path.join(folder, name)
        try:
            mtime = os.path.getmtime(full)
        except OSError:
            continue
        candidates.append((mtime, full))
    if not candidates:
        return None
    candidates.sort(reverse=True)
    return candidates[0][1]


def load_latest_routing(
    folder: str, sheet: str
) -> Tuple[Dict[Tuple[str, str], float], str, Optional[datetime]]:
    """Return (cycles_map, source_path, mtime). Empty dict on any failure."""
    src = find_latest_routing_file(folder)
    if src is None:
        return {}, "", None
    try:
        mtime = datetime.fromtimestamp(os.path.getmtime(src))
    except OSError:
        mtime = None

    try:
        wb = openpyxl.load_workbook(src, read_only=True, data_only=True)
    except Exception as e:
        logger.error("cannot open routing xlsx %s: %s", src, e)
        return {}, src, mtime

    if sheet not in wb.sheetnames:
        logger.error("sheet '%s' not in %s -- sheets: %s", sheet, src, wb.sheetnames)
        wb.close()
        return {}, src, mtime

    ws = wb[sheet]
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    wb.close()
    if not rows:
        return {}, src, mtime

    header = rows[0]
    # column 0 = "Article", columns 1+ = phase names
    phase_columns: Dict[int, str] = {}
    for idx in range(1, len(header)):
        h = header[idx]
        if h is None:
            continue
        phase_columns[idx] = str(h).strip()

    cycles: Dict[Tuple[str, str], float] = {}
    for row in rows[1:]:
        if not row:
            continue
        product = row[0]
        if product is None:
            continue
        product_code = str(product).strip()
        if not product_code:
            continue
        for col_idx, phase_name in phase_columns.items():
            if col_idx >= len(row):
                continue
            v = row[col_idx]
            if v is None:
                continue
            if isinstance(v, str):
                if v.strip().lower() == "x":
                    continue
                # try parse number
                try:
                    fv = float(v.replace(",", ".").strip())
                except ValueError:
                    logger.warning(
                        "non-numeric cell at product=%s phase=%s value=%r -- skipped",
                        product_code, phase_name, v,
                    )
                    continue
            elif isinstance(v, (int, float)):
                fv = float(v)
            else:
                logger.warning(
                    "non-numeric cell at product=%s phase=%s value=%r -- skipped",
                    product_code, phase_name, v,
                )
                continue
            if fv <= 0:
                continue
            cycles[(product_code, phase_name)] = fv

    logger.info("routing loaded: %d cycles from %s", len(cycles), src)
    return cycles, src, mtime
